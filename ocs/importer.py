"""
143/26 TP Automation
=====================
Handles:
  - Importing the transposed ST timetable (csv/xlsx)
  - Extracting per-date conducts + locations
  - Mapping conducts to SIAO lesson-plan entries and drafting the SIAO workbook
  - Building OCS/SAFTI resource (facility) bookings from the same timetable
  - Generating pre-filled Google Form URLs for facility booking submission
"""

import os
import posixpath
import zipfile
from collections import defaultdict
from datetime import datetime
from html import escape
from urllib.parse import urlencode
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from xml.etree import ElementTree as ET

import numpy as np
import openpyxl
import pandas as pd
import re
import yaml
import smtplib
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# ---------------------------------------------------------------------------
# Editable conduct catalogue
# ---------------------------------------------------------------------------

DEFAULT_CONDUCT_CATALOG_PATH = os.path.join(
    os.path.dirname(__file__),
    "conduct_catalog.yaml",
)


def normalize_conduct_name(value):
    """Normalize formatting while preserving meaningful numbers such as 4KM."""
    normalized = str(value or "").strip().upper()
    normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
    return re.sub(r"\s+", " ", normalized).strip()


def load_conduct_catalog(path=None):
    catalog_path = path or DEFAULT_CONDUCT_CATALOG_PATH
    with open(catalog_path, "r", encoding="utf-8") as catalog_file:
        catalog = yaml.safe_load(catalog_file) or {}
    catalog.setdefault("version", 1)
    catalog.setdefault("conducts", [])
    catalog.setdefault("recce_rules", [])
    return catalog


def _contains_normalized(text, term):
    normalized_text = f" {normalize_conduct_name(text)} "
    normalized_term = normalize_conduct_name(term)
    return bool(normalized_term) and f" {normalized_term} " in normalized_text


def validate_conduct_catalog(catalog, lesson_plan_names=None):
    """Return validation errors without mutating catalogue or lesson-plan data."""
    errors = []
    seen_ids = set()
    seen_aliases = defaultdict(set)
    lesson_index = {
        normalize_conduct_name(name): str(name)
        for name in (lesson_plan_names if lesson_plan_names is not None else [])
        if str(name).strip()
    }

    for row_number, conduct in enumerate(catalog.get("conducts", []), start=1):
        conduct_id = str(conduct.get("conduct_id", "")).strip()
        target = str(conduct.get("lesson_plan_name", "")).strip()
        if not conduct_id:
            errors.append(f"Conduct row {row_number} has no conduct_id.")
        elif conduct_id in seen_ids:
            errors.append(f"Duplicate conduct_id: {conduct_id}")
        seen_ids.add(conduct_id)

        if not target:
            errors.append(f"{conduct_id or f'row {row_number}'} has no lesson_plan_name.")
        elif lesson_index and normalize_conduct_name(target) not in lesson_index:
            errors.append(f"{conduct_id} targets a missing lesson-plan conduct: {target}")

        for alias in conduct.get("aliases", []):
            normalized_alias = normalize_conduct_name(alias)
            if not normalized_alias:
                errors.append(f"{conduct_id} contains a blank alias.")
                continue
            seen_aliases[normalized_alias].add(conduct_id)

    for alias, conduct_ids in seen_aliases.items():
        active_ids = {
            conduct.get("conduct_id")
            for conduct in catalog.get("conducts", [])
            if conduct.get("active", True) and conduct.get("conduct_id") in conduct_ids
        }
        if len(active_ids) > 1:
            errors.append(
                f"Alias '{alias}' is shared by active conducts: "
                + ", ".join(sorted(active_ids))
            )
    return errors


def match_catalog_conduct(conduct_text, catalog, lesson_name_index):
    """Resolve timetable text deterministically and report ambiguity explicitly."""
    normalized = normalize_conduct_name(conduct_text)
    if not normalized:
        return {"status": "unmatched", "target": None, "candidates": []}

    lesson_values = set(lesson_name_index.values())
    legacy_clean = re.sub(r"\d+", "", str(conduct_text))
    legacy_clean = re.sub(r"[^\w\s]", "", legacy_clean).rstrip().upper()
    legacy_exact = str(conduct_text).upper()
    catalog_rules = catalog.get("conducts", [])

    def rule_for_target(target):
        normalized_target = normalize_conduct_name(target)
        return next(
            (
                rule for rule in catalog_rules
                if normalize_conduct_name(rule.get("lesson_plan_name", ""))
                == normalized_target
            ),
            None,
        )

    if legacy_clean in lesson_values:
        exact_rule = rule_for_target(legacy_clean)
        if exact_rule and not exact_rule.get("active", True):
            return {
                "status": "inactive",
                "target": legacy_clean,
                "candidates": [],
            }
        return {
            "status": "exact",
            "target": legacy_clean,
            "candidates": [],
        }
    if legacy_exact in lesson_values:
        exact_rule = rule_for_target(legacy_exact)
        if exact_rule and not exact_rule.get("active", True):
            return {
                "status": "inactive",
                "target": legacy_exact,
                "candidates": [],
            }
        return {"status": "exact", "target": legacy_exact, "candidates": []}

    active_rules = [
        rule for rule in catalog_rules if rule.get("active", True)
    ]
    inactive_rules = [
        rule for rule in catalog_rules if not rule.get("active", True)
    ]
    exact_candidates = []
    contains_candidates = []

    for rule in active_rules:
        aliases = rule.get("aliases", [])
        exclusions = rule.get("exclusions", [])
        if any(_contains_normalized(conduct_text, term) for term in exclusions):
            continue
        if any(normalized == normalize_conduct_name(alias) for alias in aliases):
            exact_candidates.append(rule)
        elif any(_contains_normalized(conduct_text, alias) for alias in aliases):
            contains_candidates.append(rule)

    candidates = exact_candidates or contains_candidates
    unique_targets = {
        normalize_conduct_name(rule.get("lesson_plan_name", "")) for rule in candidates
    }
    unique_targets.discard("")

    if len(unique_targets) > 1:
        return {
            "status": "ambiguous",
            "target": None,
            "candidates": sorted(
                str(rule.get("lesson_plan_name", "")) for rule in candidates
            ),
        }
    if not candidates:
        for rule in inactive_rules:
            aliases = rule.get("aliases", [])
            exclusions = rule.get("exclusions", [])
            if any(_contains_normalized(conduct_text, term) for term in exclusions):
                continue
            if any(
                normalized == normalize_conduct_name(alias)
                or _contains_normalized(conduct_text, alias)
                for alias in aliases
            ):
                return {
                    "status": "inactive",
                    "target": str(rule.get("lesson_plan_name", "")),
                    "candidates": [],
                }
        return {"status": "unmatched", "target": None, "candidates": []}

    rule = candidates[0]
    target_key = normalize_conduct_name(rule.get("lesson_plan_name", ""))
    target = lesson_name_index.get(target_key)
    if target is None:
        return {
            "status": "invalid_target",
            "target": str(rule.get("lesson_plan_name", "")),
            "candidates": [],
        }
    return {
        "status": "catalog",
        "target": target,
        "candidates": [],
        "display_name": str(rule.get("display_name", "")).strip(),
        "use_display_name": bool(rule.get("use_display_name", False)),
    }

# ---------------------------------------------------------------------------
# Constants: SIAO / lesson-plan column mappings
# ---------------------------------------------------------------------------

SIAO_MAPPING_ARMS = {
    "PRC 650": 7,
    "PRC 940": 8,
    "DIPOLE": 9,
    "AM78": 10,
    "VRC 947": 11,
    "TA427": 12,
    "SB22": 13,
    "RL11 + D10": 14,
    "WALKIE TALKIE": 15,
    "TRS": 16,
    "BINOCULAR: 8 X 30, SWAROVSKI": 18,
    "BINOCULAR: 8 X MAGN, 30 MM": 19,
    "READER, MAP, BETALIGHT": 20,
    "COMPASS MAGNETIC": 21,
    "LASER AIMING DEVICE, VISIBLE": 22,
    "LASER AIMING DEVICE, IR": 23,
    "BORELIGHT - BL (VIS)": 24,
    "XTREK GPS": 25,
    "IN-REACH GPS": 26,
    "NVG": 27,
    "NVB": 28,
    "NIGHTHAWK": 29,
    "UTWS": 30,
    "RANGE FINDER": 31,
    "CTG 5.56MM BLANK PLASTIC A1": 44,
    "CARTRIDGE, 5.56 MM BALL, M193": 49,
    "CARTRIDGE, 5.56 MM TRACER, M196": 54,
    "CART 7.62MM M/BLANK M82 LNK M13 250R/B": 59,
    "7.62MM 4B1T,M13 LINK,250R/B": 64,
    "CTG 40MM TP S406C": 69,
    "CART 40MM ILLUM": 74,
    "CARTRIDGE,40MM HEDP-SD S408, IN M2A1 BOX": 79,
    "CART SUBCAL 18MM TRACER F/MATADOR": 84,
    "LNCHR AND CART 90MM HEAT MATADOR A1": 89,
    "5.56MM MMR RED": 94,
    "5.56MM MMR BLUE": 99,
    "CTG SIG 16MM GREEN": 104,
    "CTG SIG 16MM ILLUM": 109,
    "CTG SIG 16MM RED": 114,
    "FLARE SURFACE TRIPWIRE MK2": 119,
    "DIRECTIONAL FRAGMENTATION CHARGE M18A1": 124,
    "SIM THUNFLASH NON-ELEC": 129,
    "THUNDERFLASH,ELECTRIC": 134,
    "GRENADE HAND SMOKE SCREENING N452": 139,
    "DEMO KIT BANGALORE TORPEDO NO 21": 144,
    "BOOSTER DEMOLITION CHARGE, PETN, NO. 1": 150,
    "CHARGE DEMOLITION TNT FLAKES": 151,
    "CHGE DEMO BLK NO 3 0.5KG": 152,
    "CORD DETONATING, PETN(6000-7000 M/SEC)": 153,
    "DEMO KIT PROJECTED LINE CHGE": 154,
    "DETONATOR,NON-ELEC NO.8": 155,
    "DETONATOR ELECTRIC,NO. 8": 156,
    "FDD NO.41,W/STAINLESS STL,INNER TUBE": 157,
    "FUSE BLASTING,INSTANTANEOUS,30 M/SEC": 158,
    "FUSE BLASTING TIME SAFETY,15.24M/EA": 159,
    "IGN TIME BLAST FUSE NE": 160,
    "MATCH SAFETY 1.9 INCH": 161,
    "PLASTIC EXPLOSIVE,PETN,125GM IN M2A1 BOX": 162,
    "RKT HAND FIRED 38MM ILLUM 400M": 163,
    "RKT HAND FIRED PARACHUTE SIGNAL RED": 164,
    "SIMULATOR FLASH ARTILLERY, TNT": 165,
    "PELLET RIOT CONTROL AGENT SMOKE CS": 167,
    "SIGNAL,SMOKE:MARINE MK9": 168,
    "CARTRIDGE 9MM LUGER, SINTOX, BALL - 8G": 169,
}

ammo_types = [
    "CTG 5.56MM BLANK PLASTIC A1",
    "CARTRIDGE, 5.56 MM BALL, M193",
    "CARTRIDGE, 5.56 MM TRACER, M196",
    "CART 7.62MM M/BLANK M82 LNK M13 250R/B",
    "7.62MM 4B1T,M13 LINK,250R/B",
    "CTG 40MM TP S406C",
    "CART 40MM ILLUM",
    "CARTRIDGE,40MM HEDP-SD S408, IN M2A1 BOX",
    "CART SUBCAL 18MM TRACER F/MATADOR",
    "LNCHR AND CART 90MM HEAT MATADOR A1",
    "5.56MM MMR RED",
    "5.56MM MMR BLUE",
    "CTG SIG 16MM GREEN",
    "CTG SIG 16MM ILLUM",
    "CTG SIG 16MM RED",
    "FLARE SURFACE TRIPWIRE MK2",
    "DIRECTIONAL FRAGMENTATION CHARGE M18A1",
    "SIM THUNFLASH NON-ELEC",
    "THUNDERFLASH,ELECTRIC",
    "AMMANOL",
    "DETONATOR,NON-ELEC NO.8",
    "MATCH SAFETY 1.9 INCH",
    "FUSE BLASTING TIME SAFETY,15.24M/EA",
    "GRENADE HAND SMOKE SCREENING N452",
    "FDD NO.41,W/STAINLESS STL,INNER TUBE",
    "DEMO KIT BANGALORE TORPEDO NO 21",
]

signal_types = [
    "PRC 650",
    "PRC 940",
    "DIPOLE",
    "AM78",
    "VRC 947",
    "TA427",
    "SB22",
    "RL11 + D10",
    "WALKIE TALKIE",
    "TRS",
    "BINOCULAR: 8 X 30, SWAROVSKI",
    "BINOCULAR: 8 X MAGN, 30 MM",
    "READER, MAP, BETALIGHT",
    "COMPASS MAGNETIC",
    "LASER AIMING DEVICE, VISIBLE",
    "LASER AIMING DEVICE, IR",
    "BORELIGHT - BL (VIS)",
    "XTREK GPS",
    "IN-REACH GPS",
    "NVG",
    "NVB",
    "NIGHTHAWK",
    "UTWS",
    "RANGE FINDER",
]

SIAO_MAPPING_VEHICLES = {
    "OUV": 176,
    "SOUV": 177,
    "5-Ton": 178,
    "GP Car": 179,
    "Other Vehicles (Boats, Trailers, F550)": 180,
    "TO": 181,
    "Reporting Venue": 182,
    "Destination Venue": 183,
    "Parkover": 184,
    "Remarks": 185,
    "Indent ID": 186,
    "20 - seater": 192,
    "40 - seater": 193,
    "15 Ton Lorry": 194,
    "1 way / 2 way / Disposal": 195,
    "POC + Contact Number": 196,
    "Indent ID2": 197,
    "RPL From": 200,
    "RPL To": 201,
    "RPL Vehicle Details (with Authorised Troops)": 202,
    "Fast Craft From": 204,
    "Fast Craft To": 205,
    "Fast Craft Details (with Authorised Troops)": 206,
    "ICCT Instructors": 219,
    "SOC Key": 220,
    "Link Bridge Key": 221,
    "M1 Gate Key": 222,
    "Others": 223,
}

vehicle_types = [
    "OUV",
    "SOUV",
    "5-Ton",
    "GP Car",
    "Other Vehicles (Boats, Trailers, F550)",
    "TO",
    "Reporting Venue",
    "Destination Venue",
    "Parkover",
    "Remarks",
    "Indent ID",
    "Military Transport",
    "20 - seater",
    "40 - seater",
    "15 Ton Lorry",
    "1 way / 2 way / Disposal",
    "POC + Contact Number",
    "Indent ID2",
    "Civilian Transport",
    "RPL From",
    "RPL To",
    "RPL Vehicle Details (with Authorised Troops)",
    "Fast Craft Placeholder",
    "Fast Craft From",
    "Fast Craft To",
    "Fast Craft Details (with Authorised Troops)",
    "Sea Transport",
    "ICCT Instructors",
    "SOC Key",
    "Link Bridge Key",
    "M1 Gate Key",
    "Others",
]


SIAO_MANUAL_INPUT_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFF2CC",
)


def _has_siao_allocation(value):
    """Return whether a generated SIAO value represents a real allocation."""
    if value is None:
        return False

    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass

    if isinstance(value, str):
        return value.strip() not in {"", "-", "0", "0.0"}

    if isinstance(value, (int, float, np.number)):
        return value != 0

    return bool(value)


def highlight_siao_manual_inputs(ws, row):
    """Highlight fields that the user must complete in a generated SIAO row."""
    def fill(columns):
        for column in columns:
            ws[f"{column}{row}"].fill = SIAO_MANUAL_INPUT_FILL

    fill(("F", "H", "I", "J", "M", "N", "O", "BV", "BW"))

    ammo_base_exists = _has_siao_allocation(ws[f"T{row}"].value)
    conventional_arms_exist = any(
        _has_siao_allocation(ws.cell(row, column).value)
        for column in range(
            column_index_from_string("W"),
            column_index_from_string("AV") + 1,
        )
    )
    if ammo_base_exists or conventional_arms_exist:
        fill(("R", "S"))

    military_transport_exists = any(
        _has_siao_allocation(ws.cell(row, column).value)
        for column in range(
            column_index_from_string("CN"),
            column_index_from_string("CP") + 1,
        )
    )
    if military_transport_exists:
        fill(("CJ", "CK", "CL", "CM"))

    if _has_siao_allocation(ws[f"CX{row}"].value):
        fill(("CU", "CV", "CW"))

remarks_types = [
    "Recce",
    "Recee",
    "Predump",
]

# ---------------------------------------------------------------------------
# Constants: resource booking (facilities)
# ---------------------------------------------------------------------------

OCS_FACILITIES = {
    "CA1": ["CA1"],
    "CA2": ["CA2"],
    "SLR 1 & 2": ["SLR 1", "SLR 2"],
    "SLR 3": ["SLR 3"],
    "Company Terrain Room (CTR)": ["COMPANY TERRAIN ROOM", "CTR"],
    "Exam Hall": ["EXAM HALL"],
    "OCS Conference Room (L4)": ["OCS CONFERENCE ROOM", "L4"],
    "Interview Room": ["INTERVIEW ROOM"],
    "Cadet Mess": ["CADET MESS", "CADETS' MESS", "CADETS MESS"],
    "Multi-Purpose Hall (MPH)": ["MULTI-PURPOSE HALL", "MULTI PURPOSE HALL", "MPH"],
}

SAFTI_FACILITIES = {
    "Rambutan Hill": ["RAMBUTAN HILL", "RAMBUTAN"],
    "SOC Ground": ["SOC GROUND"],
    "Rugby Field": ["RUGBY FIELD"],
    "Warrior's Hall (WH)": ["WARRIOR'S HALL", "WARRIORS HALL", "WH"],
    "MMRC": ["MMRC"],
    "PLC IGTS": ["PLC IGTS"],
    "Stadium": ["STADIUM"],
}

FILLER_TERMS = {
    "LUNCH", "DINNER", "BREAKFAST", "-", "FREE SLOT", "HOME SWEET HOME",
}

FULL_DAY_EXCEPTIONS = [
    {
        "name": "XAW",
        "location_keywords": ["RUGBY FIELD", "MPH", "WH"],
        # at least one of these must appear in the conduct text to trigger
        "conduct_keywords": [
            "XAW", "GRASS DRILLS", "UO DRILLS",
            "HABOURING", "HARBOURING", "PTCO",
        ],
        "facilities": [
            "Rugby Field",
            "Multi-Purpose Hall (MPH)",
            "Warrior's Hall (WH)",
        ],
        "label": "XAW",
    },
]


# ---------------------------------------------------------------------------
# Module-level helpers: resource booking
# (grouped together since they're pure functions shared by Extractor.resource_booking)
# ---------------------------------------------------------------------------

def check_full_day_exception(location, conduct):
    loc_upper = location.upper()
    conduct_upper = conduct.upper()

    for rule in FULL_DAY_EXCEPTIONS:
        loc_hit = all(kw in loc_upper for kw in rule["location_keywords"])
        conduct_hit = any(kw in conduct_upper for kw in rule["conduct_keywords"])
        if loc_hit and conduct_hit:
            return rule

    return None


def time_to_minutes(t):
    h, m = t.split(":")
    return int(h) * 60 + int(m)


def minutes_to_time(m):
    return f"{m // 60:02d}:{m % 60:02d}"


def parse_time_range(time_str):
    """'0800-0850' or '0800 - 1230' -> ('08:00', '08:50')"""
    m = re.match(r"(\d{3,4})\s*-\s*(\d{3,4})", time_str.strip())
    if not m:
        return None, None

    def fmt(t):
        t = t.zfill(4)
        return f"{t[:2]}:{t[2:]}"

    return fmt(m.group(1)), fmt(m.group(2))


def day_span(time_col):
    """Global earliest start / latest end across all valid period strings."""
    starts, ends = [], []
    for val in time_col.dropna():
        s, e = parse_time_range(str(val))
        if s:
            starts.append(time_to_minutes(s))
        if e:
            ends.append(time_to_minutes(e))
    return min(starts), max(ends)


def match_facility(location, facility_map):
    loc_upper = location.upper()
    if facility_map is OCS_FACILITIES:
        slr_match = re.search(r"SLR\s*([\d,\s&/]+)", loc_upper)
        if slr_match:
            numbers = re.findall(r"\d+", slr_match.group(1))
            if "1" in numbers or "2" in numbers:
                return "SLR 1 & 2"
            if "3" in numbers:
                return "SLR 3"

    for facility, terms in facility_map.items():
        for term in terms:
            if re.search(r"\b" + re.escape(term.upper()) + r"\b", loc_upper):
                return facility

    return None


def merge_bookings(raw_bookings, day_start, day_end, max_gap_minutes=70):
    normal = []
    exception_hits = defaultdict(set)  # (date, rule_name) -> set of raw conducts seen

    for b in raw_bookings:
        rule = check_full_day_exception(b["location"], b["conduct"])
        if rule:
            exception_hits[(b["date"], rule["name"])].add(rule["name"])
        else:
            normal.append(b)

    merged = []

    # Full-day exception bookings: one row per facility, per date, full span
    for (date, rule_name), _ in exception_hits.items():
        rule = next(r for r in FULL_DAY_EXCEPTIONS if r["name"] == rule_name)
        for facility_location in rule["facilities"]:
            merged.append({
                "date": date,
                "location": facility_location,
                "start_time": minutes_to_time(day_start),
                "end_time": minutes_to_time(day_end),
                "conduct": rule["label"],
                "_skip_facility_match": True,  # already a real facility name
            })

    # --- normal contiguous/gap-bridged merge ---
    groups = defaultdict(list)
    for b in normal:
        start, end = parse_time_range(b["time"])
        if start is None:
            continue
        groups[(b["date"], b["location"])].append(
            (time_to_minutes(start), time_to_minutes(end), b["conduct"])
        )

    for (date, location), entries in groups.items():
        entries.sort(key=lambda e: e[0])
        current = None

        for start_min, end_min, conduct in entries:
            is_filler = conduct.strip().upper() in FILLER_TERMS

            if current is None:
                current = {"start": start_min, "end": end_min,
                           "conducts": [] if is_filler else [conduct]}
                continue

            gap = start_min - current["end"]
            if gap <= max_gap_minutes:
                current["end"] = max(current["end"], end_min)
                if not is_filler and conduct not in current["conducts"]:
                    current["conducts"].append(conduct)
            else:
                merged.append({
                    "date": date, "location": location,
                    "start_time": minutes_to_time(current["start"]),
                    "end_time": minutes_to_time(current["end"]),
                    "conduct": "; ".join(current["conducts"]) or conduct,
                })
                current = {"start": start_min, "end": end_min,
                           "conducts": [] if is_filler else [conduct]}

        if current:
            merged.append({
                "date": date, "location": location,
                "start_time": minutes_to_time(current["start"]),
                "end_time": minutes_to_time(current["end"]),
                "conduct": "; ".join(current["conducts"]) or "-",
            })

    return merged

def build_booking_email_html(introduction, table_html):
    """Wrap a booking table in a simple, email-safe layout."""
    paragraphs = "".join(
        f'<p style="margin:0 0 12px;">{escape(line)}</p>'
        for line in str(introduction).splitlines()
        if line.strip()
    )
    return (
        '<div style="font-family:Arial,sans-serif;color:#1f2933;font-size:14px;line-height:1.5;">'
        f"{paragraphs}{table_html}"
        "</div>"
    )


def dataframe_to_email_table(dataframe):
    """Render a dataframe as a compact HTML table suitable for email clients."""
    table = dataframe.to_html(index=False, escape=True, border=0, justify="left")
    return (
        table
        .replace(
            '<table border="0" class="dataframe">',
            '<table role="presentation" cellspacing="0" cellpadding="0" '
            'style="border-collapse:collapse;width:100%;margin-top:8px;border:1px solid #d7ded9;">',
        )
        .replace(
            "<th>",
            '<th style="background:#eef3ef;color:#18231d;text-align:left;padding:9px 10px;'
            'border:1px solid #d7ded9;white-space:nowrap;">',
        )
        .replace(
            "<td>",
            '<td style="padding:8px 10px;border:1px solid #d7ded9;vertical-align:top;">',
        )
    )


def send_email(subject, body, to_address, from_address, app_password, html=False, plain_body=None):
    """
    Sends an email via Gmail SMTP using an App Password
    (Gmail Account -> Security -> 2-Step Verification -> App Passwords).
    Credentials are read from config.yaml.
    """
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = from_address
    msg["To"] = to_address

    if html:
        if plain_body:
            msg.attach(MIMEText(plain_body, "plain", "utf-8"))
        msg.attach(MIMEText(body, "html", "utf-8"))
    else:
        msg.attach(MIMEText(body, "plain", "utf-8"))

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(from_address, app_password)
        server.sendmail(from_address, to_address, msg.as_string())

# ---------------------------------------------------------------------------
# Importer: reads raw timetable file, validates date row, forward-fills
# ---------------------------------------------------------------------------

class Importer:
    def __init__(self, path):
        if not path.endswith(('.csv', 'xlsx', 'xlsm')):
            raise ValueError("The provided file is not a Excel/CSV file.")
        self.path = path

    def import_data(self):
        read_rows = 25000
        skip_rows = 5
        if self.path.endswith(".csv"):
            return pd.read_csv(
                self.path,
                nrows=read_rows,
                skiprows=skip_rows,
                usecols=range(15),
            )

        elif self.path.endswith((".xlsx", ".xlsm")):
            data = pd.read_excel(
                self.path,
                sheet_name="ST COMBINED",
                engine="openpyxl",
                nrows=read_rows,
                skiprows=skip_rows,
                header=0,
                usecols="A:O",
            )
            return self._add_excel_text_boxes(
                data,
                sheet_name="ST COMBINED",
                skip_rows=skip_rows,
            )

    def _add_excel_text_boxes(self, data, sheet_name, skip_rows):
        """Overlay visible A:O text boxes onto their timetable lesson rows."""
        text_boxes = self._read_excel_text_boxes(sheet_name)
        if not text_boxes or data.shape[1] < 2:
            return data

        # Excel row ``skip_rows + 1`` is the pandas header, so the first
        # dataframe record begins on the following Excel row.
        first_data_row = skip_rows + 2
        lesson_rows = []
        for index, value in data.iloc[:, 1].items():
            if normalize_conduct_name(value) == "LESSON":
                lesson_rows.append((int(index), int(index) + first_data_row))

        seen = set()
        for box in text_boxes:
            excel_column = box["column"]
            if not 1 <= excel_column <= 15:
                continue

            candidates = [
                row for row in lesson_rows
                if abs(row[1] - box["row"]) <= 1
            ]
            if not candidates:
                continue

            dataframe_row, _ = min(
                candidates,
                key=lambda row: (abs(row[1] - box["row"]), row[1]),
            )
            text = box["text"].strip()
            dedupe_key = (dataframe_row, normalize_conduct_name(text))
            if not text or dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            dataframe_column = excel_column - 1
            existing = data.iat[dataframe_row, dataframe_column]
            if pd.isna(existing) or not str(existing).strip():
                data.iat[dataframe_row, dataframe_column] = text
            elif normalize_conduct_name(existing) != normalize_conduct_name(text):
                data.iat[dataframe_row, dataframe_column] = f"{existing}; {text}"

        return data

    def _read_excel_text_boxes(self, sheet_name):
        """Read DrawingML text boxes because openpyxl does not expose them."""
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        document_rel_ns = (
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        )
        package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        drawing_ns = (
            "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        )
        drawing_text_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"

        try:
            with zipfile.ZipFile(self.path) as archive:
                workbook = ET.fromstring(archive.read("xl/workbook.xml"))
                workbook_rels = ET.fromstring(
                    archive.read("xl/_rels/workbook.xml.rels")
                )
                rel_targets = {
                    rel.attrib["Id"]: rel.attrib["Target"]
                    for rel in workbook_rels.findall(
                        f"{{{package_rel_ns}}}Relationship"
                    )
                }

                sheet_path = None
                for sheet in workbook.findall(f".//{{{main_ns}}}sheet"):
                    if sheet.attrib.get("name") == sheet_name:
                        rel_id = sheet.attrib.get(f"{{{document_rel_ns}}}id")
                        target = rel_targets.get(rel_id, "")
                        sheet_path = posixpath.normpath(
                            posixpath.join("xl", target)
                        )
                        break
                if not sheet_path:
                    return []

                sheet_xml = ET.fromstring(archive.read(sheet_path))
                drawing = sheet_xml.find(f"{{{main_ns}}}drawing")
                if drawing is None:
                    return []

                sheet_rels_path = posixpath.join(
                    posixpath.dirname(sheet_path),
                    "_rels",
                    posixpath.basename(sheet_path) + ".rels",
                )
                sheet_rels = ET.fromstring(archive.read(sheet_rels_path))
                drawing_rel_id = drawing.attrib.get(f"{{{document_rel_ns}}}id")
                drawing_target = next(
                    (
                        rel.attrib.get("Target", "")
                        for rel in sheet_rels.findall(
                            f"{{{package_rel_ns}}}Relationship"
                        )
                        if rel.attrib.get("Id") == drawing_rel_id
                    ),
                    "",
                )
                if not drawing_target:
                    return []

                drawing_path = posixpath.normpath(
                    posixpath.join(posixpath.dirname(sheet_path), drawing_target)
                )
                drawing_xml = ET.fromstring(archive.read(drawing_path))

                boxes = []
                for anchor in list(drawing_xml):
                    origin = anchor.find(f"{{{drawing_ns}}}from")
                    if origin is None:
                        continue
                    row = origin.findtext(f"{{{drawing_ns}}}row")
                    column = origin.findtext(f"{{{drawing_ns}}}col")
                    if row is None or column is None:
                        continue

                    paragraphs = []
                    for paragraph in anchor.findall(
                        f".//{{{drawing_text_ns}}}p"
                    ):
                        runs = [
                            node.text or ""
                            for node in paragraph.findall(
                                f".//{{{drawing_text_ns}}}t"
                            )
                        ]
                        paragraph_text = "".join(runs).strip()
                        if paragraph_text:
                            paragraphs.append(paragraph_text)
                    text = "\n".join(paragraphs).strip()
                    if text:
                        boxes.append({
                            "row": int(row) + 1,
                            "column": int(column) + 1,
                            "text": text,
                        })
                return boxes
        except (KeyError, OSError, ValueError, zipfile.BadZipFile, ET.ParseError):
            return []

    def check_date_row(self, data: pd.DataFrame) -> list[int]:
        date_row = data.iloc[0]
        for idx, value in date_row.items():
            try:
                if isinstance(value, datetime):
                    date_row[idx] = value.strftime('%d-%b-%y')
                else:
                    parsed = pd.to_datetime(str(value), dayfirst=True, errors='raise')
                    date_row[idx] = parsed.strftime('%d-%b-%y')
            except Exception:
                # print(f"Invalid datetime format {value}")
                pass

        regex_pattern = r'\d{1,2}-[A-Za-z]{3}-\d{2}'
        has_valid_format = date_row.astype(str).str.contains(regex_pattern, regex=True).any()

        if not has_valid_format:
            raise ValueError("The first row does not contain any valid date format (DD-MMM-YY).")

        date_rows_index = pd.to_datetime(date_row, format='%d-%b-%y', errors='coerce').notna().tolist()
        index_list = [i for i, is_date in enumerate(date_rows_index) if is_date]
        return index_list

    def fill(self, data: pd.DataFrame) -> pd.DataFrame:
        filled_data = data.copy()
        filled_data.iloc[2:13] = data.iloc[2:13].ffill(axis=0)
        return filled_data


# ---------------------------------------------------------------------------
# Extractor: turns the filled timetable into conduct/location tables,
# maps conducts to SIAO lesson-plan entries, drafts the SIAO workbook,
# and builds OCS/SAFTI resource bookings.
# ---------------------------------------------------------------------------

class Extractor:
    def __init__(self, data: pd.DataFrame, extract_columns=None, lesson_plan_path=None,
                 siao_template_path=None, conduct_catalog_path=None):
        self.data = data
        self.extract_columns = extract_columns
        self.lesson_plan_path = lesson_plan_path
        self.siao_template_path = siao_template_path
        self.conduct_catalog_path = conduct_catalog_path or DEFAULT_CONDUCT_CATALOG_PATH
        self.conduct_catalog = load_conduct_catalog(self.conduct_catalog_path)
        self.match_report = []

    def format_date_header(self, value):
        try:
            return pd.to_datetime(value, dayfirst=True).strftime("%d-%b-%y")
        except:
            return str(value)

    def extract(self):
        dc = pd.DataFrame()
        for col in self.extract_columns:
            try:
                header = self.format_date_header(self.data.columns[col])

                if not re.search(r'\d{1,2}-[A-Za-z]{3}-\d{2}', header):
                    continue

                conduct = self.data.iloc[:, col - 1]
                location = self.data.iloc[:, col + 1].ffill()
                conduct = conduct.reindex(location.index)
                conduct = conduct[conduct.index != "REMARKS"]
                location = location[location.index != "REMARKS"]
                dc[col] = conduct
                dc[col + 1] = location

            except (IndexError, re.error, TypeError):
                print(
                    f"Column '{self.data.columns[col]}' not found in data "
                    "or contains invalid regex pattern. Skipping."
                )

        dc.columns = [
            x
            for col in self.extract_columns
            for x in (
                self.format_date_header(self.data.columns[col]),
                "Location"
            )
        ]
        self.conducts = dc
        print(self.conducts)

    def indents(self, conduct_mapping):
        pass

    def read_lesson_plan(self):
        lesson_plan = pd.read_csv(self.lesson_plan_path)
        header = lesson_plan.iloc[0:1].copy()
        header = header.ffill(axis=1)
        lesson_plan.columns = header.iloc[0]
        lesson_plan = lesson_plan.iloc[1:].reset_index(drop=True)
        self.lesson_plan = lesson_plan
        lesson_names = lesson_plan.iloc[2:, 2].dropna().astype(str)
        self.lesson_name_index = {}
        self.lesson_name_conflicts = []
        for lesson_name in lesson_names:
            normalized = normalize_conduct_name(lesson_name)
            if normalized in self.lesson_name_index:
                self.lesson_name_conflicts.append(normalized)
            else:
                self.lesson_name_index[normalized] = lesson_name
        self.catalog_validation_errors = validate_conduct_catalog(
            self.conduct_catalog,
            lesson_names,
        )

    def draft_siao(self, cadet_size: int):
        self.extract()
        self.read_lesson_plan()
        conduct_mapping = self.conduct_list()
        remarks = data_change.remarks_information()
        recce = data_change.recce_information(remarks)
        for date, lines in recce.items():
            if date in conduct_mapping:
                conduct_mapping[date].extend([lines[0], lines[1]])
            else:
                conduct_mapping[date] = [lines[0], lines[1]]
        conduct_mapping = self.conduct_manipulation(conduct_mapping)
        conduct_mapping = dict(
            sorted(
                conduct_mapping.items(),
                key=lambda item: datetime.strptime(
                    item[0].split(",")[0].strip(),
                    "%d-%b-%y"
                )
            )
        )
        # print(conduct_mapping)

        template_path = self.siao_template_path

        wb = openpyxl.load_workbook(template_path)
        ws = wb["(Fill In) SIAO"]

        start_row = 13
        processed = set()

        for c in conduct_mapping:
            for match_name, display_name in zip(
                conduct_mapping[c][0::2],
                conduct_mapping[c][1::2]
            ):
                if (display_name, c) in processed:
                    continue
                processed.add((display_name, c))

                matches = self.siao_conducts[self.siao_conducts == match_name]

                if matches.empty:
                    continue

                row_num = matches.index[0]
                l = []

                try:
                    medic = (
                        "Y"
                        if int(self.lesson_plan.iloc[row_num, 37]) >= 1
                        else "-"
                    )
                except:
                    medic = "-"

                if len(c.split(',')) > 1:
                    start_date = c.split(',')[0].strip()
                    end_date = c.split(',')[-1].strip()
                else:
                    start_date = c
                    end_date = c
                # Conduct details + Medic
                l.extend([
                    display_name,
                    "",
                    start_date,
                    "",
                    end_date,
                    "",
                    "",
                    "",
                    self.lesson_plan.iloc[row_num, 3],
                    "",
                    "",
                    "",
                    "",
                    self.lesson_plan.iloc[row_num, 37],
                    medic,
                    "",
                    "",
                    self.lesson_plan.iloc[row_num, 41],
                    "",
                    "-"
                ])

                for ammo in ammo_types:
                    if ammo == "AMMANOL":
                        l.append(0)
                        continue

                    l.append(
                        int(self.lesson_plan.iloc[row_num, SIAO_MAPPING_ARMS[ammo]])
                        * cadet_size
                    )

                # Demolition and others
                l.extend(["-"] * 25)

                # Start / End Time (TODO)
                l.extend(["-", "-"])

                # Vehicles
                for vehicle in vehicle_types:
                    if vehicle == "Military Transport":
                        l.extend(["-", "-", "-", "-", "-"])
                        continue

                    elif vehicle == "Civilian Transport":
                        l.extend(["-", "-"])
                        continue

                    elif vehicle == "Fast Craft Placeholder":
                        l.append("-")
                        continue

                    elif vehicle == "Sea Transport":
                        l.extend(["-"] * 12)
                        continue

                    l.append(
                        self.lesson_plan.iloc[
                            row_num,
                            SIAO_MAPPING_VEHICLES[vehicle]
                        ]
                    )

                # Others End Placeholder
                l.append("-")

                # Write row into Excel starting from column C
                for col_num, value in enumerate(l, start=3):
                    ws.cell(row=start_row, column=col_num, value=value)

                highlight_siao_manual_inputs(ws, start_row)

                start_row += 1

        wb.save(template_path)

    def conduct_list(self):
        conduct_list = self.conducts
        lesson_plan = self.lesson_plan

        self.siao_conducts = lesson_plan.iloc[2:, 2]
        date_conduct_mapping = {}
        cl = {}
        self.match_report = []

        for i, column in enumerate(conduct_list.columns):
            date = (
                f"{conduct_list.columns[i-1]}_Location"
                if column == "Location"
                else column
            )

            conducts = []

            for conduct in conduct_list.iloc[:, i].dropna():
                if conduct == "":
                    continue

                conducts.extend(
                    c.strip()
                    for c in str(conduct).split(";")
                    if c.strip()
                )

            cl[date] = conducts

        for date, conducts in cl.items():
            for conduct in conducts:
                if conduct == "LOC":
                    break

                resolution = None
                if match := re.search(r"\bSOC\s+\d+\b", str(conduct), re.IGNORECASE):
                    target = self.lesson_name_index.get(
                        normalize_conduct_name(match.group(0))
                    )
                    if target:
                        resolution = {
                            "status": "exact",
                            "target": target,
                            "candidates": [],
                        }
                if resolution is None:
                    resolution = match_catalog_conduct(
                        conduct,
                        self.conduct_catalog,
                        self.lesson_name_index,
                    )

                self.match_report.append({
                    "date": date,
                    "conduct": str(conduct),
                    "status": resolution["status"],
                    "lesson_plan_name": resolution.get("target") or "",
                    "candidates": "; ".join(resolution.get("candidates", [])),
                })
                if resolution["status"] in ("exact", "catalog"):
                    display_name = conduct
                    if resolution.get("use_display_name") and resolution.get("display_name"):
                        display_name = resolution["display_name"]
                    date_conduct_mapping.setdefault(date, []).extend(
                        [resolution["target"], display_name]
                    )

        return date_conduct_mapping

    def conduct_manipulation(self, conduct_mapping):
        exercise_ranges = self.conduct_exercises(conduct_mapping)

        new_mapping = {}
        exercise_entries = {}

        for date, conducts in conduct_mapping.items():
            remaining = []

            for match_name, display_name in zip(conducts[0::2], conducts[1::2]):

                if match_name in exercise_ranges:
                    if match_name not in exercise_entries:
                        rule = self.catalog_rule_for_target(match_name)
                        if rule and rule.get("exercise_display_name"):
                            display_name = rule["exercise_display_name"]

                        exercise_entries[match_name] = [
                            match_name,
                            display_name
                        ]

                    continue

                remaining.extend([match_name, display_name])

            if remaining:
                new_mapping[date] = remaining

        for match_name, pair in exercise_entries.items():
            new_key = ", ".join(exercise_ranges[match_name])
            new_mapping[new_key] = pair

        return new_mapping

    def catalog_rule_for_target(self, target):
        normalized_target = normalize_conduct_name(target)
        for rule in self.conduct_catalog.get("conducts", []):
            if normalize_conduct_name(rule.get("lesson_plan_name", "")) == normalized_target:
                return rule
        return None

    def conduct_exercises(self, conduct_mapping):
        exercise = {}
        for c in conduct_mapping:
            for match_name in conduct_mapping[c][0::2]:
                rule = self.catalog_rule_for_target(match_name)
                if rule and rule.get("multi_day", False):
                    if match_name not in exercise:
                        exercise[match_name] = c
                    else:
                        exercise[match_name] += f", {c}"

        for k, v in exercise.items():
            dates = [d.strip() for d in v.split(",")]
            exercise[k] = [dates[0], dates[-1]]
        return exercise

    def remarks_information(self):
        data = self.data.copy()

        remarks = {}

        for col in self.extract_columns:
            try:
                header = self.format_date_header(data.columns[col])

                if not re.search(r"\d{1,2}-[A-Za-z]{3}-\d{2}", header):
                    continue

                indices = [i for i in (col, col - 1, col - 2) if i >= 0]

                if any(pd.notna(data.iloc[13, i]) for i in indices):
                    remark = next(data.iloc[13, i] for i in indices if pd.notna(data.iloc[13, i]))
                    remarks[header] = remark

            except (IndexError, re.error, TypeError):
                continue

        return remarks

    def recce_information(self, remarks):
        recce = {}

        for date, remark in remarks.items():
            if not isinstance(remark, str):
                continue

            if not any(rt.lower() in remark.lower() for rt in remarks_types):
                continue

            for line in remark.splitlines():
                line = line.strip()
                if not line:
                    continue

                for rule in self.conduct_catalog.get("recce_rules", []):
                    aliases = rule.get("aliases", [])
                    if any(_contains_normalized(line, alias) for alias in aliases):
                        configured_target = rule.get("lesson_plan_name", "")
                        target = self.lesson_name_index.get(
                            normalize_conduct_name(configured_target),
                            configured_target,
                        )
                        recce[date] = [target, line]
                        break

        return recce

    def resource_booking_details(self):
        if not hasattr(self, "conducts"):
            self.extract()

        conducts = self.conducts
        columns = conducts.columns
        time_col = self.data.iloc[:, 0]

        booking = []
        i = 0
        n = len(columns)

        while i < n:
            column = columns[i]

            if column == "Location":
                i += 1
                continue

            date = column

            if i + 1 >= n or columns[i + 1] != "Location":
                i += 1
                continue

            conduct_series = conducts.iloc[:, i]
            location_series = conducts.iloc[:, i + 1]

            for row_idx in conduct_series.index:
                conduct = conduct_series.loc[row_idx]
                location = location_series.loc[row_idx]

                if pd.isna(conduct) or str(conduct).strip() in ("", "-"):
                    continue
                if pd.isna(location) or str(location).strip() in ("", "-"):
                    continue

                try:
                    time_value = time_col.loc[row_idx]
                except KeyError:
                    time_value = None

                if pd.isna(time_value) or str(time_value).strip() == "":
                    continue

                time_str = str(time_value).strip()

                for single_conduct in str(conduct).split(";"):
                    single_conduct = single_conduct.strip()
                    if not single_conduct:
                        continue

                    booking.append({
                        "date": date,
                        "time": time_str,
                        "location": str(location).strip(),
                        "conduct": single_conduct,
                    })

            i += 2
        return booking

    def resource_booking(self, your_email, output_path="ocs_booking.csv",
                          gmail_address=None, gmail_app_password=None):
        raw = self.resource_booking_details()

        time_col = self.data.iloc[:, 0]
        d_start, d_end = day_span(time_col)

        merged = merge_bookings(raw, d_start, d_end)

        ocs_rows, safti_rows = [], []

        for b in merged:
            location = b["location"]
            row_template = {
                "START DATE": b["date"],
                "START TIME": b["start_time"],
                "END DATE": b["date"],
                "END TIME": b["end_time"],
                "REMARKS/JUSTIFICATIONS": b["conduct"],
                "YOUR EMAIL": your_email,
            }

            if b.get("_skip_facility_match"):
                # already resolved to a real facility name in FULL_DAY_EXCEPTIONS
                if location in OCS_FACILITIES:
                    ocs_rows.append({"FACILITY": location, **row_template})
                elif location in SAFTI_FACILITIES:
                    safti_rows.append({"FACILITY": location, **row_template})
                continue

            ocs_facility = match_facility(location, OCS_FACILITIES)
            if ocs_facility:
                ocs_rows.append({"FACILITY": ocs_facility, **row_template})

            safti_facility = match_facility(location, SAFTI_FACILITIES)
            if safti_facility:
                safti_rows.append({"FACILITY": safti_facility, **row_template})

        # 1. OCS facilities -> CSV (pandas) + printed copy/paste block
        headers = [
            "FACILITY", "START DATE", "START TIME",
            "END DATE", "END TIME", "REMARKS/JUSTIFICATIONS", "YOUR EMAIL",
        ]
        ocs_df = pd.DataFrame(ocs_rows, columns=headers)
        ocs_df.to_csv(output_path, index=False)

        print(f"--- OCS Facility Bookings ({len(ocs_df)}) -> {output_path} ---")
        print(ocs_df.to_string(index=False))

        # 2. SAFTI facilities -> draft email (not sent yet)
        safti_df = pd.DataFrame(safti_rows, columns=headers)

        email_intro = "Requesting the following SAFTI facility bookings:"
        table_text = safti_df.to_string(index=False)
        table_html = dataframe_to_email_table(safti_df)
        email_draft = {
            "to": your_email,
            "subject": "SAFTI Facility Booking Request",
            "body": email_intro,
            "table_text": table_text,
            "table_html": table_html,
            "plain_body": f"{email_intro}\n\n{table_text}",
            "html_body": build_booking_email_html(email_intro, table_html),
        }

        print("\n--- SAFTI Email Draft (not sent) ---")
        print(f"To: {email_draft['to']}")
        print(f"Subject: {email_draft['subject']}")
        print(email_draft["plain_body"])

        if gmail_address and gmail_app_password:
            send_email(
                subject=email_draft["subject"],
                body=email_draft["html_body"],
                to_address=email_draft["to"],
                from_address=gmail_address,
                app_password=gmail_app_password,
                html=True,
                plain_body=email_draft["plain_body"],
            )
            print(f"\n--- SAFTI Email sent to {email_draft['to']} ---")
        else:
            print("\n--- SAFTI Email Draft (not sent - no Gmail credentials configured) ---")
            print(f"To: {email_draft['to']}")
            print(f"Subject: {email_draft['subject']}")
            print(email_draft["plain_body"])

        return {"ocs": ocs_df, "safti": safti_df, "email_draft": email_draft}


# ---------------------------------------------------------------------------
# GoogleFormSubmitter: builds pre-filled Google Form URLs from config.yaml
# ---------------------------------------------------------------------------

class GoogleFormSubmitter:
    def __init__(self, config):
        self.form_url = config["google_form"]["url"]
        self.form_fields = config["google_form"]["fields"]

    def generate_prefilled_url(self, data):

        params = {}

        for field_name, field_info in self.form_fields.items():

            field_id = field_info["id"]
            field_type = field_info.get("type", "text")

            value = data.get(field_name, field_info.get("default", ""))

            if value == "":
                continue

            if field_type == "time":
                value = value.replace(":", "")

                if len(value) == 3:
                    value = "0" + value

                hour = value[:2]
                minute = value[2:]

                params[f"{field_id}_hour"] = hour
                params[f"{field_id}_minute"] = minute

            else:
                params[field_id] = value

        return f"{self.form_url}?{urlencode(params)}"


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    config_path = os.path.join(os.path.dirname(__file__), "config.yaml")
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)

    paths = config["paths"]

    importer = Importer(paths["input_data"])
    data = importer.import_data()
    data_transposed = data.transpose()
    data_transposed.columns = data_transposed.iloc[0]
    # data_transposed.to_csv(r"transposed_data_test2.csv", index=False)
    l_extract_list = importer.check_date_row(data_transposed)
    data_transposed = data_transposed.drop(data_transposed.index[0])
    # data_transposed.to_csv(r"transposed_data_test3.csv", index=False)
    functional_data = importer.fill(data_transposed)
    # functional_data.to_csv(r"transposed_data_test4.csv", index=False)

    data_change = Extractor(
        functional_data,
        extract_columns=l_extract_list,
        lesson_plan_path=paths["lesson_plan"],
        siao_template_path=paths["siao_template"],
    )
    data_change.draft_siao(cadet_size=120)

    resource_booking_output = os.path.join(paths["output_folder"], "ocs_booking.csv")
    gmail_config = config.get("gmail", {})
    data_change.resource_booking(
        your_email=config["user"]["email"],
        output_path=resource_booking_output,
        gmail_address=gmail_config.get("address") or None,
        gmail_app_password=gmail_config.get("app_password") or None,
    )

    # g = GoogleFormSubmitter(config)
    # google_data = {
    # "facility": "Cadet Mess",
    # "start_date": "2026-07-07",
    # "end_date": "2026-07-07",
    # "start_time": "0900",
    # "end_time": "1100",
    # "rank": "PTE",
    # "name": "Kai Quan",
    # "contact_number": "91234567"
    # }

    # print(g.generate_prefilled_url(google_data))
