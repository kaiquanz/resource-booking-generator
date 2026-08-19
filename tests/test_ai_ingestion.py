import io
import contextlib
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pandas as pd
import yaml
from pypdf import PdfWriter

from ai_ingestion import (
    DEFAULT_AI_MODEL,
    EVENT_COLUMNS,
    TrainingPlanExtraction,
    _date_span_from_texts,
    build_ai_input,
    extract_training_plan_with_ai,
    validate_reviewed_events,
)
from app_services import (
    apply_configured_period_ranges,
    build_booking_email_content,
    canonical_events_to_functional_data,
    generate_bookings_from_events,
    generate_siao_from_events,
)


ROOT = Path(__file__).resolve().parents[1]


def event_frame(rows):
    return pd.DataFrame(rows, columns=EVENT_COLUMNS)


class FakeResponses:
    def __init__(self):
        self.kwargs = None

    def parse(self, **kwargs):
        self.kwargs = kwargs
        parsed = TrainingPlanExtraction.model_validate({
            "document_title": "Mock TP",
            "events": [{
                "date": "2026-10-01",
                "start_time": "08:00",
                "end_time": "09:00",
                "conduct": "STRENGTH TRAINING",
                "location": "CA1",
                "remarks": "",
                "source_reference": "page 1, box A",
                "confidence": 0.96,
                "needs_review": False,
            }],
            "warnings": [],
        })
        return SimpleNamespace(
            status="completed",
            output_parsed=parsed,
            output=[],
            model=DEFAULT_AI_MODEL,
            id="resp_test",
        )


class FakeClient:
    def __init__(self):
        self.responses = FakeResponses()


class ChunkedFakeResponses:
    def __init__(self):
        self.calls = []

    def parse(self, **kwargs):
        self.calls.append(kwargs)
        call_number = len(self.calls)
        dates = ["2026-10-01", "2026-10-15", "2026-11-05"]
        parsed = TrainingPlanExtraction.model_validate({
            "document_title": "Long TP",
            "events": [{
                "date": dates[call_number - 1],
                "start_time": "08:00",
                "end_time": "09:00",
                "conduct": f"EVENT {call_number}",
                "location": "Training Area",
                "remarks": "",
                "source_reference": f"original page {call_number * 2 - 1}",
                "confidence": 0.95,
                "needs_review": False,
            }],
            "warnings": [],
        })
        return SimpleNamespace(
            status="completed",
            output_parsed=parsed,
            output=[],
            model=DEFAULT_AI_MODEL,
            id=f"resp_chunk_{call_number}",
        )


class ChunkedFakeClient:
    def __init__(self):
        self.responses = ChunkedFakeResponses()


class AIIngestionTests(unittest.TestCase):
    def test_visible_pdf_date_span_includes_the_final_month(self):
        start_date, end_date = _date_span_from_texts([
            "Week 1: 1-Oct-26 to 7-Oct-26",
            "Final week: 24-Nov-26 to 30-Nov-26",
        ])

        self.assertEqual(start_date, "2026-10-01")
        self.assertEqual(end_date, "2026-11-30")

    def test_pdf_uses_ephemeral_file_input(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "visual-plan.pdf"
            source.write_bytes(b"%PDF-1.4 test")
            content = build_ai_input(source)

        self.assertEqual(content[0]["type"], "input_file")
        self.assertEqual(content[0]["filename"], "visual-plan.pdf")
        self.assertEqual(content[0]["detail"], "high")
        self.assertTrue(content[0]["file_data"].startswith("data:application/pdf;base64,"))
        self.assertEqual(content[1]["type"], "input_text")

    def test_tsv_uses_an_accepted_mime_type(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "plan.tsv"
            source.write_text("date\tconduct\n", encoding="utf-8")
            content = build_ai_input(source)

        self.assertTrue(content[0]["file_data"].startswith("data:text/tsv;base64,"))

    def test_xlsm_is_converted_to_supported_xlsx_input(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "plan.xlsm"
            workbook = openpyxl.Workbook()
            workbook.active["A1"] = "Training plan"
            workbook.save(source)
            workbook.close()
            content = build_ai_input(source)

        self.assertEqual(content[0]["filename"], "plan.xlsx")
        self.assertTrue(
            content[0]["file_data"].startswith(
                "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
            )
        )

    def test_image_uses_original_detail(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "boxed-plan.png"
            source.write_bytes(b"not-a-real-image")
            content = build_ai_input(source)

        self.assertEqual(content[0]["type"], "input_image")
        self.assertEqual(content[0]["detail"], "original")
        self.assertTrue(content[0]["image_url"].startswith("data:image/png;base64,"))

    def test_structured_extraction_request_and_result(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "plan.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=792, height=612)
            with source.open("wb") as stream:
                writer.write(stream)
            client = FakeClient()
            result = extract_training_plan_with_ai(
                source,
                api_key="test-key",
                client=client,
            )

        request = client.responses.kwargs
        self.assertEqual(DEFAULT_AI_MODEL, "gpt-5.6-luna")
        self.assertEqual(request["model"], DEFAULT_AI_MODEL)
        self.assertEqual(request["max_output_tokens"], 64_000)
        self.assertIs(request["text_format"], TrainingPlanExtraction)
        self.assertFalse(request["store"])
        self.assertEqual(result["model"], DEFAULT_AI_MODEL)
        self.assertEqual(result["events"].loc[0, "conduct"], "STRENGTH TRAINING")
        self.assertEqual(result["coverage_label"], "1/1 PDF pages")

    def test_configured_period_zero_is_added_to_ai_prompt(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "plan.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=792, height=612)
            with source.open("wb") as stream:
                writer.write(stream)
            client = FakeClient()
            extract_training_plan_with_ai(
                source,
                api_key="test-key",
                client=client,
                period_definitions={
                    "0": {"start_time": "07:00", "end_time": "07:50"}
                },
            )

        prompt = client.responses.kwargs["input"][0]["content"][1]["text"]
        self.assertIn("Period 0: 07:00-07:50", prompt)
        self.assertIn("only when the source names that period", prompt)

    def test_configured_period_zero_fills_a_missing_time_range(self):
        data = pd.DataFrame(
            {"TIME": [None, "0800-0850"]},
            index=["Period 0", "Period 1"],
        )
        config = {
            "timetable": {
                "periods": {
                    "0": {"start_time": "07:00", "end_time": "07:50"}
                }
            }
        }

        filled = apply_configured_period_ranges(data, config)

        self.assertEqual(filled.loc["Period 0", "TIME"], "0700-0750")
        self.assertEqual(filled.loc["Period 1", "TIME"], "0800-0850")

    def test_pdf_chunks_reach_and_merge_the_final_month(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "long-plan.pdf"
            writer = PdfWriter()
            for _ in range(5):
                writer.add_blank_page(width=792, height=612)
            with source.open("wb") as stream:
                writer.write(stream)
            client = ChunkedFakeClient()
            result = extract_training_plan_with_ai(
                source,
                api_key="test-key",
                client=client,
            )

        self.assertEqual(len(client.responses.calls), 3)
        self.assertEqual(result["page_ranges"], ["1-2", "3-4", "5"])
        self.assertEqual(result["coverage_label"], "5/5 PDF pages")
        self.assertEqual(result["chunks_processed"], 3)
        self.assertEqual(result["events"]["date"].tolist()[-1], "2026-11-05")
        self.assertTrue(
            all(call["model"] == DEFAULT_AI_MODEL for call in client.responses.calls)
        )
        final_prompt = client.responses.calls[-1]["input"][0]["content"][1]["text"]
        self.assertIn("original page 5", final_prompt)
        self.assertIn("5-page source PDF", final_prompt)

    def test_review_validation_preserves_iso_dates(self):
        reviewed = event_frame([{
            "date": "2026-10-01",
            "start_time": "800",
            "end_time": "09:15",
            "conduct": "STRENGTH TRAINING",
            "location": "CA1",
            "remarks": "",
            "source_reference": "page 1",
            "confidence": 0.9,
            "needs_review": "false",
        }])
        cleaned, errors, warnings = validate_reviewed_events(reviewed)

        self.assertEqual(errors, [])
        self.assertEqual(warnings, [])
        self.assertEqual(cleaned.loc[0, "date"], "2026-10-01")
        self.assertEqual(cleaned.loc[0, "start_time"], "08:00")
        self.assertFalse(bool(cleaned.loc[0, "needs_review"]))

    def test_adapter_preserves_legacy_positions(self):
        reviewed = event_frame([
            {
                "date": "2026-10-01", "start_time": "08:00", "end_time": "09:00",
                "conduct": "STRENGTH TRAINING", "location": "CA1", "remarks": "First note",
                "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
            },
            {
                "date": "2026-10-02", "start_time": "10:00", "end_time": "11:00",
                "conduct": "IPPT", "location": "", "remarks": "Second note",
                "source_reference": "page 2", "confidence": 1.0, "needs_review": False,
            },
        ])
        functional, date_columns = canonical_events_to_functional_data(reviewed)

        self.assertEqual(date_columns, [2, 5])
        self.assertEqual(functional.columns[2], "01-Oct-26")
        self.assertEqual(functional.columns[5], "02-Oct-26")
        self.assertEqual(functional.index[13], "REMARKS")
        self.assertEqual(functional.iloc[0, 3], "LOC")
        self.assertEqual(functional.iloc[0, 6], "LOC")
        self.assertEqual(functional.iloc[1, 0], "0800-0900")
        self.assertEqual(functional.iloc[2, 6], "-")
        self.assertEqual(functional.iloc[13, 1], "First note")
        self.assertEqual(functional.iloc[13, 4], "Second note")

    def test_ai_events_preserve_the_automation_workbook_shape(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        reviewed = event_frame([{
            "date": "2026-10-01", "start_time": "08:00", "end_time": "09:00",
            "conduct": "IPPT", "location": "CA1", "remarks": "",
            "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
        }])
        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_siao_from_events(config, reviewed, cadet_size=120)

        template = openpyxl.load_workbook(ROOT / "ocs" / "template_siao.xlsx")
        roundtrip_bytes = io.BytesIO()
        template.save(roundtrip_bytes)
        roundtrip_bytes.seek(0)
        automation_baseline = openpyxl.load_workbook(roundtrip_bytes)
        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]))
        self.assertEqual(generated.sheetnames, automation_baseline.sheetnames)
        for sheet_name in automation_baseline.sheetnames:
            expected = automation_baseline[sheet_name]
            actual = generated[sheet_name]
            self.assertEqual(actual.max_row, expected.max_row)
            self.assertEqual(actual.max_column, expected.max_column)
            self.assertEqual(
                list(actual.merged_cells.ranges),
                list(expected.merged_cells.ranges),
            )
        template.close()
        automation_baseline.close()
        generated.close()
        self.assertIn("exact", set(result["match_report"]["status"]))

    def test_siao_uses_catalogued_preparation_windows_and_new_vehicle_columns(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        reviewed = event_frame([{
            "date": "2026-10-12", "start_time": "08:00", "end_time": "09:00",
            "conduct": "Ex Adaptive Warrior ( CO + UO)",
            "location": "Rugby Field", "remarks": "",
            "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
        }])

        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_siao_from_events(config, reviewed, cadet_size=120)

        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]), data_only=False)
        sheet = generated["(Fill In) SIAO"]
        self.assertEqual(sheet["M13"].value, "12/Oct/2026 08:00")
        self.assertEqual(sheet["N13"].value, "13/Oct/2026 21:00")
        self.assertEqual(sheet["R13"].value, "12/Oct/2026 05:30")
        self.assertEqual(sheet["S13"].value, "13/Oct/2026 22:00")
        self.assertEqual(sheet["BQ13"].value, "12/Oct/2026 05:30")
        self.assertEqual(sheet["BR13"].value, "13/Oct/2026 22:00")
        self.assertEqual(sheet["CG13"].value, "-")
        self.assertEqual(sheet["CH13"].value, "-")
        self.assertEqual(sheet["BS13"].value, "-")
        self.assertEqual(sheet["BT13"].value, "-")
        self.assertEqual(sheet["BU13"].value, "1")
        generated.close()

    def test_siao_handles_recce_row_without_catalogue_rule(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        reviewed = event_frame([{
            "date": "2026-10-12", "start_time": "08:00", "end_time": "09:00",
            "conduct": "Ex Adaptive Warrior ( CO + UO)",
            "location": "Rugby Field", "remarks": "XAW Recce @ Rambutan Hill",
            "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
        }])

        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_siao_from_events(config, reviewed, cadet_size=120)

        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]), data_only=False)
        sheet = generated["(Fill In) SIAO"]
        self.assertEqual(sheet["C13"].value, "XAW Recce @ Rambutan Hill")
        self.assertEqual(sheet["C14"].value, "Ex Adaptive Warrior ( CO + UO)")
        self.assertEqual(sheet["E14"].value, "12-Oct-26")
        self.assertEqual(sheet["G14"].value, "13-Oct-26")
        self.assertEqual(sheet["CG13"].value, "-")
        self.assertEqual(sheet["CH13"].value, "-")
        generated.close()

    def test_exact_lesson_plan_match_uses_catalogued_bus_setting(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        reviewed = event_frame([{
            "date": "2026-10-07", "start_time": "07:00", "end_time": "07:50",
            "conduct": "Signal Package", "location": "Stagmont Camp", "remarks": "",
            "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
        }])

        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_siao_from_events(config, reviewed, cadet_size=140)

        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]), data_only=False)
        sheet = generated["(Fill In) SIAO"]
        self.assertEqual(sheet["C13"].value, "Signal Package")
        self.assertEqual(sheet["CI13"].value, "SAFTI MI")
        self.assertEqual(sheet["CJ13"].value, "STAGMONT CAMP")
        self.assertEqual(sheet["CL13"].value, 4)
        generated.close()

        catalog = yaml.safe_load(
            (ROOT / "ocs" / "conduct_catalog.yaml").read_text(encoding="utf-8")
        )
        signal_rule = next(
            rule for rule in catalog["conducts"]
            if rule["conduct_id"] == "signal_package"
        )
        signal_rule["bus_required"] = False
        with tempfile.TemporaryDirectory() as temp_dir:
            catalog_path = Path(temp_dir) / "conduct_catalog.yaml"
            catalog_path.write_text(
                yaml.safe_dump(catalog, sort_keys=False),
                encoding="utf-8",
            )
            config["paths"]["conduct_catalog"] = str(catalog_path)
            with contextlib.redirect_stdout(io.StringIO()):
                result = generate_siao_from_events(config, reviewed, cadet_size=140)

        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]), data_only=False)
        self.assertEqual(generated["(Fill In) SIAO"]["CL13"].value, "-")
        generated.close()

    def test_transport_window_populates_general_and_military_transport_fields(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        reviewed = event_frame([{
            "date": "2026-10-23", "start_time": "06:15", "end_time": "16:00",
            "conduct": "M203 L/F", "location": "M203 RANGE", "remarks": "",
            "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
        }])

        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_siao_from_events(config, reviewed, cadet_size=140)

        generated = openpyxl.load_workbook(io.BytesIO(result["xlsx"]), data_only=False)
        sheet = generated["(Fill In) SIAO"]
        self.assertEqual(sheet["BQ13"].value, "23/Oct/2026 04:45")
        self.assertEqual(sheet["BR13"].value, "23/Oct/2026 22:30")
        self.assertEqual(sheet["CG13"].value, "23/Oct/2026 06:15")
        self.assertEqual(sheet["CH13"].value, "23/Oct/2026 16:00")
        self.assertEqual(sheet["CI13"].value, "Tango Wing")
        self.assertEqual(sheet["CJ13"].value, "M203 Range")
        self.assertEqual(sheet["CL13"].value, 4)
        generated.close()

    def test_ai_events_generate_ocs_and_safti_products(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        config["user"]["email"] = "reviewer@example.com"
        reviewed = event_frame([
            {
                "date": "2026-10-01", "start_time": "08:00", "end_time": "09:00",
                "conduct": "STRENGTH TRAINING", "location": "CA1", "remarks": "",
                "source_reference": "page 1", "confidence": 1.0, "needs_review": False,
            },
            {
                "date": "2026-10-02", "start_time": "10:00", "end_time": "11:00",
                "conduct": "IPPT", "location": "Stadium", "remarks": "",
                "source_reference": "page 2", "confidence": 1.0, "needs_review": False,
            },
        ])
        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_bookings_from_events(config, reviewed)

        self.assertEqual(result["ocs"].loc[0, "FACILITY"], "CA1")
        self.assertEqual(result["safti"].loc[0, "FACILITY"], "Stadium")
        self.assertEqual(result["email_draft"]["to"], "reviewer@example.com")
        self.assertIn("<table", result["email_draft"]["table_html"])

    def test_ai_bookings_without_recipient_create_copyable_email_draft(self):
        config = yaml.safe_load((ROOT / "ocs" / "config.yaml").read_text(encoding="utf-8"))
        config["user"]["email"] = ""
        reviewed = event_frame([{
            "date": "2026-10-02", "start_time": "10:00", "end_time": "11:00",
            "conduct": "IPPT", "location": "Stadium", "remarks": "",
            "source_reference": "page 2", "confidence": 1.0, "needs_review": False,
        }])

        with contextlib.redirect_stdout(io.StringIO()):
            result = generate_bookings_from_events(config, reviewed)

        self.assertEqual(result["email_draft"]["to"], "")
        self.assertIn("Subject: SAFTI Facility Booking Request", result["email_copy_text"])
        self.assertIn("FACILITY", result["email_copy_text"])
        self.assertIn("Stadium", result["email_copy_text"])
        content = build_booking_email_content(result["email_draft"])
        self.assertIn("<table", content["html"])
        self.assertIn("Stadium", content["html"])


if __name__ == "__main__":
    unittest.main()
